import os

from openpyxl import Workbook
from openpyxl import load_workbook

from .base_excel_man import BaseExcelMan


class ExcelManOpenPyXl(BaseExcelMan):
    def __init__(self, file_path, sheet_name=None, has_title=True):
        super().__init__(file_path, sheet_name=None, has_title=True)

    def initialization(self, file_path, sheet_name=None):
        if os.path.exists(file_path):
            workbook = load_workbook(filename=file_path)
            if sheet_name is None:
                worksheet = workbook[workbook.sheetnames[0]]
            else:
                worksheet = workbook[sheet_name]
        else:
            workbook = Workbook()
            worksheet = workbook.active
            if sheet_name is None:
                worksheet.title = 'Sheet1'
            else:
                worksheet.title = sheet_name
        return workbook, worksheet

    def get_lines_and_columns(self, line_or_col=None, title=None):
        values = []
        if title is None:
            line_or_col = line_or_col
        else:
            line_or_col = self.titles[title]
        for val_obj in self.worksheet[line_or_col]:
            values.append(val_obj.value)
        return values

    def get_lines_and_columns_num(self,line_or_col):
        return len(self.get_lines_and_columns(line_or_col))

    def get_title(self):
        '''
        :return:
        '''
        title_line = self.get_lines_and_columns(1)
        for i in range(0, len(title_line)):
            k = title_line[i]
            if i <= 25:
                v = self.char_list['A-Z'][i]

            else:
                index = int(i / 26) - 1
                first_letter = self.char_list['A-Z'][index]
                second_letter = self.char_list['A-Z'][i - 26 * (index + 1)]
                v = first_letter + second_letter
            self.titles[k] = v

    def write_row(self, row_value_list, istitle=False):
        if istitle:
            self.write_spicified_row(specified_row=1, row_value_list=row_value_list)
        else:
            self.worksheet.append(row_value_list)
        self.workbook.save(self.file_path)

    def write_spicified_row(self, specified_row, row_value_list):
        self.worksheet.insert_rows(specified_row)
        for i in range(0, len(row_value_list)):
            if i <= 25:
                self.worksheet[self.char_list['A-Z'][i] + str(specified_row)] = row_value_list[i]
            else:
                index = int(i / 26) - 1
                first_letter = self.char_list['A-Z'][index]
                second_letter = self.char_list['A-Z'][i - 26]
                self.worksheet[first_letter + second_letter + str(specified_row)] = row_value_list[i]
        self.workbook.save(self.file_path)

    def del_specified_rows(self, specified_rows, del_blank_rows=False):
        specified_rows.sort(reverse = False)
        for i in specified_rows:
            self.worksheet.delete_rows(i)
            self.worksheet.insert_rows(i)
        if del_blank_rows:
            for i in  specified_rows:
                self.worksheet.delete_rows(i-specified_rows.index(i))
        self.workbook.save(self.file_path)

    def del_specified_cols(self, specified_cols=None, titles=None, del_blank_cols=False):
        if specified_cols is None and titles is None:
            raise Exception('must specify one of specified_cols or titles')
        if specified_cols is None:
            specified_cols = []
            for col in titles:
                specified_cols.append(self.titles[col])
        specified_cols.sort(reverse=False)
        for col in specified_cols:
            title = self.worksheet[col+str(1)].value
            self.worksheet.delete_cols(self.char_list['A-Z'].index(col)+1)
            self.worksheet.insert_cols(self.char_list['A-Z'].index(col)+1)
            self.worksheet[col+str(1)] = title
        self.workbook.save(self.file_path)

    def write_specified_grid(self,  value, line, title=None, col_name=None):
        if title is None and col_name is None:
            raise Exception('must specify one of col_name or titles')
        if title is None:
            grid = col_name + str(line)
        else:
            grid = self.titles[title] + str(line)
        self.worksheet[grid] = str(value)
        self.workbook.save(self.file_path)